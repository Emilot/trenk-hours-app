import os
import threading
import time
from queue import Queue
import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime, time as dtime
from openpyxl.utils import column_index_from_string
from report_logic import generate_monthly_report
from utils.spreadsheet_utils import get_column_from_day, open_excel

INVALID_TIME_VALUES = [
    None, "", "0", "null", "#null", "#NULL",
    "#TIMH!", "#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#N/A"
]

def parse_hours_range(text):
    """Δέχεται string 'HH:MM-HH:MM' και επιστρέφει διάρκεια σε ώρες (float)."""
    if not text or not isinstance(text, str):
        return None
    s = text.strip().upper()
    if s in ("", "ΡΕΠΟ"):
        return None
    try:
        start_str, end_str = s.split("-")
        fmt = "%H:%M"
        start = datetime.strptime(start_str.strip(), fmt)
        end = datetime.strptime(end_str.strip(), fmt)
        if end < start:
            end = end.replace(day=end.day + 1)
        duration = (end - start).total_seconds() / 3600.0
        return round(duration, 3)
    except Exception:
        return None

def _format_time_cell(v):
    if isinstance(v, dtime):
        return v.strftime("%H:%M")
    if isinstance(v, datetime):
        return v.strftime("%H:%M")
    if isinstance(v, (int, float)):
        # Excel stores times as fraction of day
        frac = float(v) % 1.0
        total_minutes = int(round(frac * 24 * 60))
        hh = (total_minutes // 60) % 24
        mm = total_minutes % 60
        return f"{hh:02d}:{mm:02d}"
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%H:%M", "%H:%M:%S"):
            try:
                dt = datetime.strptime(s, fmt)
                return dt.strftime("%H:%M")
            except Exception:
                pass
        if len(s) >= 5 and s[2] == ":":
            return s[:5]
        return ""
    return ""

def update_cell(ws, cell_name, value):
    """Robust A1 -> (row, col) update using openpyxl cell by index."""
    if not isinstance(cell_name, str) or not cell_name:
        raise ValueError("Άκυρο cell_name")
    col_letters = ''.join(filter(str.isalpha, cell_name))
    row_digits = ''.join(filter(str.isdigit, cell_name))
    if not col_letters or not row_digits:
        raise ValueError(f"Άκυρη διεύθυνση κελιού: {cell_name}")
    row_number = int(row_digits)
    col_index = column_index_from_string(col_letters)
    ws.cell(row=row_number, column=col_index, value=value)

def _get_sheet(wb, candidates):
    # Prefer exact names, but try stripped names too
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    stripped = {s.strip(): s for s in wb.sheetnames}
    for name in candidates:
        key = name.strip()
        if key in stripped:
            return wb[stripped[key]]
    raise KeyError(f"Δεν βρέθηκε κανένα από τα φύλλα: {candidates}")

def _format_seconds(secs):
    secs = max(0, int(secs))
    if secs < 60:
        return f"{secs}s"
    m, s = divmod(secs, 60)
    if m < 60:
        return f"{m}m {s}s"
    h, m = divmod(m, 60)
    return f"{h}h {m}m"

def main():
    global root, txt_output, btn_open_excel, month_selector, controls

    root = tk.Tk()
    root.title("Trenkwalder Payroll Tool")
    root.geometry("900x650")

    weekly_file = tk.StringVar()
    payroll_file = tk.StringVar()
    selected_month = tk.IntVar(value=7)

    txt_output = tk.Text(root, wrap="word", height=16)

    def browse_weekly():
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if path:
            weekly_file.set(path)

    def browse_payroll():
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if path:
            payroll_file.set(path)

    # --- GUI Components ---
    tk.Label(root, text="Εβδομαδιαίο αρχείο:").pack(anchor="w", padx=10, pady=(12, 2))
    tk.Entry(root, textvariable=weekly_file, width=90).pack(fill="x", padx=10)
    btn_browse_weekly = tk.Button(root, text="Browse", command=browse_weekly)
    btn_browse_weekly.pack(anchor="w", padx=10, pady=(2, 10))

    tk.Label(root, text="Payroll αρχείο:").pack(anchor="w", padx=10, pady=(0, 2))
    tk.Entry(root, textvariable=payroll_file, width=90).pack(fill="x", padx=10)
    btn_browse_payroll = tk.Button(root, text="Browse", command=browse_payroll)
    btn_browse_payroll.pack(anchor="w", padx=10, pady=(2, 10))

    tk.Label(root, text="Μήνας υπολογισμού (1–12):").pack(anchor="w", padx=10)
    month_selector = ttk.Combobox(
        root, textvariable=selected_month,
        values=list(range(1, 13)), state="readonly", width=10
    )
    month_selector.pack(anchor="w", padx=10, pady=(2, 10))

    btn_run = tk.Button(root, text="Υπολογισμός")
    btn_run.pack(pady=6)

    btn_open_excel = tk.Button(
        root, text="Άνοιγμα Excel",
        command=lambda: open_excel(
            os.path.join(os.path.dirname(payroll_file.get()), "Payroll_Calculated.xlsx")
        )
    )
    btn_open_excel.pack(pady=(2, 10))
    btn_open_excel.config(state="disabled")

    txt_output.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    # Λίστα για κλείδωμα/ξεκλείδωμα controls
    controls = [btn_browse_weekly, btn_browse_payroll, btn_run, btn_open_excel, month_selector]

    # --- Progress communication (worker -> UI) ---
    progress_q = Queue()

    # --- Loader Overlay (υβριδικό progress) ---
    loader_overlay = None
    loader_bar = None
    loader_stage = None
    eta_label = None

    loader_running = False
    start_time_parse = 0.0
    last_ui_update_t = 0.0

    progress_state = {"value": 0, "stage": "idle"}

    report_fake_active = False
    report_fake_target = 95
    report_fake_timer_id = None

    def start_loader(text="Επεξεργασία..."):
        nonlocal loader_overlay, loader_bar, loader_stage, eta_label
        nonlocal loader_running, start_time_parse, last_ui_update_t
        nonlocal report_fake_active, report_fake_timer_id

        for w in controls:
            if w is month_selector:
                w.config(state="disabled")
            else:
                w.config(state="disabled")

        loader_overlay = tk.Frame(root, bg="#000000", highlightthickness=0)
        loader_overlay.place(relx=0, rely=0, relwidth=1, relheight=1)

        box = tk.Frame(loader_overlay, bg="white", padx=22, pady=18, bd=1, relief="solid")
        box.place(relx=0.5, rely=0.5, anchor="center")

        loader_stage = tk.Label(box, text=text, bg="white", font=("Segoe UI", 10))
        loader_stage.pack(pady=(0, 8))

        loader_bar = ttk.Progressbar(box, mode="determinate", length=360, maximum=100)
        loader_bar.pack()
        loader_bar["value"] = 0

        eta_label = tk.Label(box, text="Προετοιμασία...", bg="white", fg="#666", font=("Segoe UI", 9))
        eta_label.pack(pady=(8, 0))

        loader_running = True
        start_time_parse = 0.0
        last_ui_update_t = 0.0
        progress_state["value"] = 0
        progress_state["stage"] = "parse"
        report_fake_active = False
        report_fake_timer_id = None

        root.after(80, _poll_queue)

    def stop_loader():
        nonlocal loader_overlay, loader_bar, loader_stage, eta_label
        nonlocal loader_running, report_fake_active, report_fake_timer_id

        loader_running = False

        if report_fake_timer_id is not None:
            try:
                root.after_cancel(report_fake_timer_id)
            except Exception:
                pass
            report_fake_timer_id = None
        report_fake_active = False

        try:
            if loader_overlay:
                loader_overlay.place_forget()
                loader_overlay.destroy()
        finally:
            loader_overlay = None
            loader_bar = None
            loader_stage = None
            eta_label = None
            for w in controls:
                if w is btn_open_excel:
                    continue
                if w is month_selector:
                    w.config(state="readonly")
                else:
                    w.config(state="normal")

    def _start_report_fake_fill():
        nonlocal report_fake_active, report_fake_timer_id
        report_fake_active = True

        def _tick():
            nonlocal report_fake_timer_id
            if not loader_running or not report_fake_active or loader_bar is None:
                return
            val = progress_state["value"]
            if val < report_fake_target:
                step = max(0.2, (report_fake_target - val) * 0.06)
                progress_state["value"] = min(report_fake_target, val + step)
                loader_bar["value"] = progress_state["value"]
            report_fake_timer_id = root.after(180, _tick)

        report_fake_timer_id = root.after(180, _tick)

    def _stop_report_fake_fill():
        nonlocal report_fake_active, report_fake_timer_id
        report_fake_active = False
        if report_fake_timer_id is not None:
            try:
                root.after_cancel(report_fake_timer_id)
            except Exception:
                pass
            report_fake_timer_id = None

    def _poll_queue():
        nonlocal last_ui_update_t, start_time_parse
        if not loader_running:
            return

        changed = False
        latest_text = None
        stage_changed = False

        # Drain queue; keep only latest state updates and immediately inject logs
        while not progress_q.empty():
            msg = progress_q.get_nowait()
            mtype = msg.get("type")
            if mtype == "stage":
                name = msg.get("name", "")
                text = msg.get("text", "")
                if name and name != progress_state["stage"]:
                    progress_state["stage"] = name
                    stage_changed = True
                latest_text = text or latest_text
                changed = True
            elif mtype == "set_val":
                try:
                    v = float(msg.get("val", progress_state["value"]))
                except Exception:
                    v = progress_state["value"]
                progress_state["value"] = max(0.0, min(100.0, v))
                changed = True
            elif mtype == "inc":
                try:
                    dv = float(msg.get("by", 1.0))
                except Exception:
                    dv = 1.0
                progress_state["value"] = max(0.0, min(100.0, progress_state["value"] + dv))
                changed = True
            elif mtype == "log":
                txt_output.insert("end", msg.get("msg", "") + "\n")
                txt_output.see("end")

        if stage_changed and loader_stage is not None:
            if progress_state["stage"] == "parse":
                _stop_report_fake_fill()
                start_time_parse = 0.0
                loader_stage.config(text="Ανάλυση δεδομένων...")
                eta_label.config(text="Υπολογισμός εκτιμ. χρόνου...")
            elif progress_state["stage"] == "report":
                loader_stage.config(text="Υπολογισμός μισθοδοσίας...")
                eta_label.config(text="Υπολογισμός... (εκτίμηση)")
                _start_report_fake_fill()
            elif progress_state["stage"] == "save":
                _stop_report_fake_fill()
                loader_stage.config(text="Αποθήκευση αρχείου...")
                eta_label.config(text="Σχεδόν έτοιμο...")

        now = time.time()
        if changed and (now - last_ui_update_t) >= 0.1 and loader_bar is not None:
            loader_bar["value"] = progress_state["value"]

            if progress_state["stage"] == "parse":
                if progress_state["value"] > 0:
                    if start_time_parse == 0.0:
                        start_time_parse = now
                    elapsed = now - start_time_parse
                    processed_pct = min(0.999, progress_state["value"] / 80.0) if progress_state["value"] <= 80 else 1.0
                    if processed_pct > 0:
                        total_est = elapsed / processed_pct
                        remaining = max(0.0, total_est - elapsed)
                        eta_label.config(text=f"Πρόοδος: {progress_state['value']:.0f}% — Εκτιμ. υπόλοιπο: {_format_seconds(remaining)}")
                    else:
                        eta_label.config(text="Υπολογισμός εκτιμ. χρόνου...")
                else:
                    eta_label.config(text="Έναρξη...")
            elif progress_state["stage"] == "report":
                eta_label.config(text="Υπολογισμός... (εκτίμηση)")
            elif progress_state["stage"] == "save":
                eta_label.config(text="Ολοκλήρωση...")

            last_ui_update_t = now

        if loader_running:
            root.after(100, _poll_queue)

    def run_export():
        try:
            txt_output.delete("1.0", tk.END)
            weekly_path = weekly_file.get().strip()
            payroll_path = payroll_file.get().strip()
            month = selected_month.get()

            if not (1 <= month <= 12):
                raise ValueError("Ο μήνας πρέπει να είναι μεταξύ 1 και 12.")
            if not weekly_path or not payroll_path:
                raise ValueError("Πρέπει να επιλέξετε και τα δύο αρχεία.")
            if not weekly_path.endswith(".xlsx") or not payroll_path.endswith(".xlsx"):
                raise ValueError("Τα αρχεία πρέπει να είναι τύπου .xlsx")

            start_loader("Ανάλυση δεδομένων...")

            thread = threading.Thread(
                target=_export_task,
                args=(weekly_path, payroll_path, month, progress_q),
                daemon=True
            )
            thread.start()

        except Exception as e:
            messagebox.showerror("Σφάλμα", str(e))

    def _export_task(weekly_path, payroll_path, month, q: Queue):
        """
        Hybrid progress:
        - Parsing known size => 0–80%
        - Report unknown => fake fill 80–95%
        - Save => 95–100%
        """
        try:
            q.put({"type": "stage", "name": "parse", "text": "Ανάλυση δεδομένων..."})
            q.put({"type": "set_val", "val": 0})

            # load weekly schedule in data_only mode for safe reads
            wb_weekly = openpyxl.load_workbook(weekly_path, data_only=True)
            sheet_weekly = _get_sheet(wb_weekly, ["ΦΟΡΜΑ ΚΑΤΑΧΩΡΙΣΗΣ ", "ΦΟΡΜΑ ΚΑΤΑΧΩΡΙΣΗΣ"])
            sheet_times = _get_sheet(wb_weekly, ["ΥΠΕΡΕΡΓΑΣΙΕΣ-ΥΠΕΡΩΡΙΕΣ"])

            DAY_TO_COLS = {
                0: ('C', 'D'), 1: ('H', 'I'), 2: ('M', 'N'),
                3: ('R', 'S'), 4: ('W', 'X'), 5: ('AB', 'AC'), 6: ('AG', 'AH'),
            }

            # Efficient two-pass: first pass counts potential entries quickly (no full parse),
            # second pass builds schedule_rows. Counting is lightweight: check for non-empty cells.
            total_entries = 0
            min_r = 10
            max_col_for_count = 9  # we only check first 9 columns as before
            for row in sheet_weekly.iter_rows(min_row=min_r, max_col=max_col_for_count, values_only=True):
                if not row:
                    continue
                # hours columns are columns 3..9 in 1-based -> indices 2..8
                for val in row[2:9]:
                    if val is None:
                        continue
                    s = str(val).strip()
                    if not s:
                        continue
                    # cheap heuristic: contains '-' or ':' likely a range
                    if "-" in s or ":" in s:
                        total_entries += 1

            total_entries = max(1, total_entries)
            tick_every = max(1, total_entries // 80)

            schedule_rows = []
            skipped_entries = []
            done_entries = 0

            # Precompute column index cache for DAY_TO_COLS letters
            col_index_cache = {}
            for letters in DAY_TO_COLS.values():
                for letter in letters:
                    if letter not in col_index_cache:
                        col_index_cache[letter] = column_index_from_string(letter)

            # Second pass: build rows
            for idx, row in enumerate(sheet_weekly.iter_rows(min_row=min_r, max_col=max_col_for_count, values_only=True), start=min_r):
                try:
                    full_id = str(row[0]).strip() if row and row[0] else ""
                    work_type = str(row[1]).strip() if row and row[1] else ""
                    hours_list = row[2:9] if row else ()
                    afm = full_id.split()[0] if full_id else ""

                    if not afm:
                        continue

                    for i, hours_raw in enumerate(hours_list):
                        if hours_raw is None or str(hours_raw).strip() == "":
                            continue

                        hours_value = parse_hours_range(str(hours_raw))
                        if hours_value is None:
                            continue

                        date_cell = sheet_weekly.cell(row=8, column=3 + i)
                        date_raw = date_cell.value
                        if not isinstance(date_raw, datetime):
                            continue

                        dow = date_raw.weekday()
                        letters = DAY_TO_COLS.get(dow)
                        if not letters:
                            continue
                        left_letter, right_letter = letters
                        left_col = col_index_cache[left_letter]
                        right_col = col_index_cache[right_letter]

                        raw_end_plus_30 = sheet_times.cell(row=idx, column=left_col).value
                        raw_departure = sheet_times.cell(row=idx, column=right_col).value

                        entry_end_plus_30 = _format_time_cell(raw_end_plus_30)
                        entry_departure = _format_time_cell(raw_departure)

                        schedule_rows.append({
                            "date": date_raw,
                            "employee": afm,
                            "hours": hours_value,
                            "work_type": work_type,
                            "ΩΡΑΡΙΟ": hours_value,
                            "ΩΡΑ ΛΗΞΗΣ+30": entry_end_plus_30,
                            "ΩΡΑ ΑΠΟΧΩΡΗΣΗ": entry_departure,
                        })

                        done_entries += 1
                        if done_entries % tick_every == 0:
                            mapped = min(80, int(done_entries * 80 / total_entries))
                            q.put({"type": "set_val", "val": mapped})

                except Exception as err:
                    skipped_entries.append(f"γραμμή {idx} ➤ {row[0] if row else ''} - ΣΦΑΛΜΑ: {str(err)}")

            q.put({"type": "set_val", "val": 80})

            q.put({"type": "stage", "name": "report", "text": "Υπολογισμός μισθοδοσίας..."})

            wb_payroll = openpyxl.load_workbook(payroll_path)
            sheet_payroll = _get_sheet(wb_payroll, ["ΩΡΟΜΕΤΡΗΣΗ"])

            # GUI wrapper: route messages to queue
            class SpreadsheetWrapper:
                def __init__(self, ws):
                    self.ws = ws
                    self.wb = wb_payroll
                def update_cell(self, cell_name, value):
                    # keep original semantics: update_cell(ws, a1, value)
                    update_cell(self.ws, cell_name, value)

            class GUIWrapper:
                def show_message(self, msg, level="info"):
                    q.put({"type": "log", "msg": msg})

            spreadsheet = SpreadsheetWrapper(sheet_payroll)
            gui = GUIWrapper()

            generate_monthly_report(
                schedule_rows, month, spreadsheet, gui,
                get_column_from_day, overtime_ws=sheet_times,
                forma_wb=wb_weekly
            )

            q.put({"type": "stage", "name": "save", "text": "Αποθήκευση αρχείου..."})

            save_path = os.path.join(os.path.dirname(payroll_path), "Payroll_Calculated.xlsx")
            wb_payroll.save(save_path)

            q.put({"type": "set_val", "val": 100})

            root.after(0, lambda: _finish_export({
                "success": True,
                "save_path": save_path,
                "skipped_entries": skipped_entries,
            }))

        except Exception as e:
            root.after(0, lambda: _finish_export({
                "success": False,
                "error": str(e),
            }))

    def _finish_export(result):
        stop_loader()

        if result.get("success"):
            txt_output.insert("end", f"\n✅ Αποθήκευση στο: {result['save_path']}\n")
            txt_output.see("end")
            btn_open_excel.config(state="normal")
            if result.get("skipped_entries"):
                txt_output.insert("end", "\n⚠️ Παραλείφθηκαν εγγραφές με σφάλματα:\n")
                txt_output.insert("end", "\n".join(result["skipped_entries"]) + "\n")
        else:
            messagebox.showerror("Σφάλμα", f"Προέκυψε σφάλμα:\n{result['error']}")
            txt_output.insert("end", f"❌ Σφάλμα:\n{result['error']}\n")
            txt_output.see("end")

    btn_run.config(command=run_export)

    root.mainloop()

if __name__ == "__main__":
    main()