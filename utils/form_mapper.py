from openpyxl.utils import get_column_letter, column_index_from_string
from utils.overtime_utils import DAY_TO_COLS_INT, day_names, ot_columns
import datetime as _dt
from typing import Tuple, Dict, Set, Any


def _parse_cell_date(cell_val: Any) -> _dt.date | None:
    """
    Try to coerce various cell values to a date (datetime.date).
    Accepts:
      - datetime.date / datetime.datetime
      - strings in common formats like 'dd/mm/YYYY', 'dd/mm/yy', 'YYYY-mm-dd'
    Returns None if parsing is not possible (we intentionally do not try
    to decode Excel serial numbers here).
    """
    if cell_val is None:
        return None
    # Already a date/datetime
    if isinstance(cell_val, _dt.datetime):
        return cell_val.date()
    if isinstance(cell_val, _dt.date):
        return cell_val
    # Strings: try several formats
    s = str(cell_val).strip()
    if not s:
        return None
    fmts = ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d")
    for fmt in fmts:
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except Exception:
            continue
    return None


def build_day_map(ws_form) -> Tuple[Dict[int, Dict[str, str]], Set[int]]:
    """
    Build a mapping for the 7-day block in the form (columns C..I) that maps:
      day_of_month -> {
        "form_col": <A1 column for that day in the form sheet>,
        "ot_end_col": <column letter in overtime sheet for end+30>,
        "ot_leave_col": <column letter in overtime sheet for departure>
      }
    Also returns a set of days that correspond to Sundays in that block.

    Notes / improvements over the prior implementation:
    - Parses both datetime and common string formats robustly.
    - Uses weekday -> DAY_TO_COLS_INT mapping (from utils.overtime_utils) instead
      of relying solely on enumeration order, which is safer if column ranges change.
    - Skips cells that cannot be parsed to a date (rather than raising).
    - Stops mixing month boundaries: it establishes the initial_month from the
      first valid date and ignores subsequent cells that belong to a different month.
    """
    day_map: Dict[int, Dict[str, str]] = {}
    sunday_days: Set[int] = set()
    initial_month = None

    start_col = column_index_from_string("C")
    end_col = column_index_from_string("I")

    for col_idx in range(start_col, end_col + 1):
        cell_val = ws_form.cell(row=8, column=col_idx).value
        dt_date = _parse_cell_date(cell_val)
        if not dt_date:
            continue

        # Ensure we stay in the same month as the first found date in the block
        if initial_month is None:
            initial_month = dt_date.month
        elif dt_date.month != initial_month:
            # skip days that belong to a different month
            continue

        day = dt_date.day
        weekday = dt_date.weekday()  # 0=Monday .. 6=Sunday

        # get overtime columns either via numeric mapping or fall back to name mapping
        ot_pair = DAY_TO_COLS_INT.get(weekday)
        if not ot_pair:
            # fallback: use day_names list to map to ot_columns (keeps backward compatibility)
            try:
                ot_pair = ot_columns[day_names[weekday]]
            except Exception:
                # if neither mapping exists, skip this day
                continue

        ot_end_col, ot_leave_col = ot_pair

        day_map[day] = {
            "form_col": get_column_letter(col_idx),
            "ot_end_col": ot_end_col,
            "ot_leave_col": ot_leave_col
        }

        if weekday == 6:  # Sunday
            sunday_days.add(day)

    return day_map, sunday_days