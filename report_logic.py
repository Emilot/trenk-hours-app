from utils.metrics import get_metric_rows, inspect_sunday_metrics, update_sundays
from calendar import monthrange
from datetime import datetime, date, time, timedelta
from openpyxl.utils import column_index_from_string, get_column_letter
from collections import defaultdict

INVALID_TIME_VALUES = [
    None, "", "0", "null", "#null", "#NULL",
    "#TIMH!", "#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#N/A"
]
INVALID_TIME_TOKENS = {str(v).strip().upper() for v in INVALID_TIME_VALUES if isinstance(v, str)}

DAY_TO_COLS = {
    0: ('C',  'D'),
    1: ('H',  'I'),
    2: ('M',  'N'),
    3: ('R',  'S'),
    4: ('W',  'X'),
    5: ('AB', 'AC'),
    6: ('AG', 'AH'),
}

AFM_COL_ΩΡΟΜΕΤΡΗΣΗ = 5  # Ε
MAX_ROW = 200

def clean_time_string(time_str):
    if isinstance(time_str, str):
        return time_str.replace(".", ":").strip()
    return ""

def is_valid_time_string(value):
    if value is None:
        return False
    if isinstance(value, (datetime, time)):
        return True
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        s = value.strip()
        if s == "":
            return False
        if s.upper() in INVALID_TIME_TOKENS:
            return False
        for fmt in ("%H:%M", "%H:%M:%S", "%H:%M:%S.%f"):
            try:
                datetime.strptime(s.replace(".", ":"), fmt)
                return True
            except Exception:
                pass
        return False
    return False

def _to_hhmm(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.replace(".", ":").strip()
    if isinstance(value, time):
        return f"{value.hour:02d}:{value.minute:02d}"
    if isinstance(value, datetime):
        return f"{value.hour:02d}:{value.minute:02d}"
    if isinstance(value, (int, float)):
        total_minutes = int(round(float(value) * 24 * 60))
        total_minutes = max(0, total_minutes % (24 * 60))
        hh = total_minutes // 60
        mm = total_minutes % 60
        return f"{hh:02d}:{mm:02d}"
    s = str(value).strip()
    return s.replace(".", ":")

def read_work_times_from_sheet(ws_source, anchor_row_idx: int, day_date: date, gui=None) -> dict:
    dow = day_date.weekday()
    left_col_letter, right_col_letter = DAY_TO_COLS[dow]
    left_col_idx = column_index_from_string(left_col_letter)
    right_col_idx = column_index_from_string(right_col_letter)

    left_val_raw = ws_source.cell(row=anchor_row_idx, column=left_col_idx).value
    right_val_raw = ws_source.cell(row=anchor_row_idx, column=right_col_idx).value

    left_val = _to_hhmm(left_val_raw)
    right_val = _to_hhmm(right_val_raw)

    debug_msg = (
        f"🧾 Κελί χρόνου ({ws_source.title}) ➤ Ημέρα: {day_date.strftime('%A %d/%m')}\n"
        f"🔹 {left_col_letter}{anchor_row_idx} ➤ raw='{left_val_raw}' | τύπος={type(left_val_raw).__name__} → καθαρό='{left_val}'\n"
        f"🔹 {right_col_letter}{anchor_row_idx} ➤ raw='{right_val_raw}' | τύπος={type(right_val_raw).__name__} → καθαρό='{right_val}'"
    )
    gui.show_message(debug_msg, level="debug") if gui else None

    return {
        "ΩΡΑ ΛΗΞΗΣ+30": left_val,
        "ΩΡΑ ΑΠΟΧΩΡΗΣΗ": right_val,
        "_cells": (f"{left_col_letter}{anchor_row_idx}", f"{right_col_letter}{anchor_row_idx}")
    }

def calculate_overtime(end_plus_30, departure_time, date_obj):
    fmt = "%H:%M"
    try:
        start = datetime.strptime(end_plus_30, fmt).time()
        end = datetime.strptime(departure_time, fmt).time()
    except Exception:
        return {"ΥΠΕΡΕΡΓΑΣΙΑ": 0, "ΥΠΕΡΩΡΙΑ": 0, "ΑΡΓΙΑ": 0}

    start_dt = datetime.combine(date_obj, start)
    end_dt = datetime.combine(date_obj, end)
    if end_dt <= start_dt:
        end_dt += timedelta(days=1)

    diff_minutes = (end_dt - start_dt).total_seconds() / 60
    is_sunday = date_obj.weekday() == 6

    if diff_minutes <= 0:
        return {"ΥΠΕΡΕΡΓΑΣΙΑ": 0, "ΥΠΕΡΩΡΙΑ": 0, "ΑΡΓΙΑ": 0}

    if diff_minutes <= 60:
        yperergasia = round(diff_minutes / 60, 2)
        yperoria = 0
    else:
        yperergasia = 1.0
        yperoria = round((diff_minutes - 60) / 60, 2)

    argia = round(yperergasia + yperoria, 2) if is_sunday else 0

    return {"ΥΠΕΡΕΡΓΑΣΙΑ": yperergasia, "ΥΠΕΡΩΡΙΑ": yperoria, "ΑΡΓΙΑ": argia}

def calculate_night_hours(start_str: str, end_str: str) -> float:
    fmt = "%H:%M"
    try:
        start_dt = datetime.strptime(start_str, fmt)
        end_dt = datetime.strptime(end_str, fmt)
        if end_dt <= start_dt:
            end_dt += timedelta(days=1)
        night_start = start_dt.replace(hour=22, minute=0, second=0, microsecond=0)
        night_end = night_start + timedelta(hours=8)
        overlap_start = max(start_dt, night_start)
        overlap_end = min(end_dt, night_end)
        if overlap_end <= overlap_start:
            return 0.0
        minutes = (overlap_end - overlap_start).total_seconds() / 60.0
        return round(minutes / 60.0, 3)
    except Exception:
        return 0.0

def update_excel_cell(ws, cell_name, value):
    col_letters = ''.join(filter(str.isalpha, cell_name))
    row_number = int(''.join(filter(str.isdigit, cell_name)))
    col_index = column_index_from_string(col_letters)
    ws.cell(row=row_number, column=col_index).value = value

def normalize_afm_strict(val: str) -> str:
    digits = "".join(ch for ch in str(val) if ch.isdigit())
    return digits.zfill(9) if digits else ""

def is_valid_afm(s):
    s = str(s).strip()
    return s.isdigit() and len(s) == 9

def find_employee_row_in_sheet(ws, afm, gui=None, diagnostics=False, *,
                               min_row=1, max_row=None,
                               strict_cell_match=False,
                               search_columns=None,
                               cache=None):
    """
    Returns list of 1-based row indices where the AFM appears in the given worksheet.
    If cache (dict) is provided, it must be used as a mapping keyed by (id(ws), normalized_afm)
    so that cached results are worksheet-specific.

    This avoids cross-sheet cache pollution (bug fixed from earlier refactor).
    """
    # Build a sheet-aware cache key when a cache dict is provided
    key = None
    if cache is not None:
        key = (id(ws), str(afm).strip())
        cached = cache.get(key)
        if cached is not None:
            return list(cached)

    target_afm = normalize_afm_strict(afm)
    matches = []
    try:
        max_row_eff = max_row or ws.max_row
        col_range = list(search_columns) if search_columns else None

        for idx in range(min_row, max_row_eff + 1):
            values = [ws.cell(row=idx, column=c).value for c in col_range] if col_range else [cell.value for cell in ws[idx]]
            for val in values:
                if val is None:
                    continue
                cell_afm = normalize_afm_strict(val)
                if strict_cell_match:
                    if cell_afm == target_afm:
                        if diagnostics:
                            msg = f"🔎 Βρέθηκε ΑΦΜ {target_afm} (exact) στη γραμμή {idx} του φύλλου '{ws.title}'"
                            gui.show_message(msg, level="debug") if gui else None
                        matches.append(idx)
                        break
                else:
                    if (cell_afm and target_afm in cell_afm) or (target_afm and target_afm in str(val)):
                        if diagnostics:
                            msg = f"🔎 Βρέθηκε ΑΦΜ {target_afm} ως substring στη γραμμή {idx} του φύλλου '{ws.title}'"
                            gui.show_message(msg, level="debug") if gui else None
                        matches.append(idx)
                        break

        if cache is not None:
            cache[key] = list(matches)

        if diagnostics and not matches:
            msg = f"⚠️ Δεν βρέθηκε ΑΦΜ {target_afm} στο φύλλο '{ws.title}'"
            gui.show_message(msg, level="warning") if gui else None
        return matches
    except Exception as ex:
        error_msg = f"❌ Σφάλμα στη find_employee_row_in_sheet για ΑΦΜ {target_afm}: {str(ex)}"
        gui.show_message(error_msg, level="error") if gui else None
        return []

def normalize_label(s: str) -> str:
    if not isinstance(s, str):
        return ""
    return str(s).strip().upper().replace(" ", "")

def compute_anchor(row: int) -> int:
    return row - ((row - 2) % 6)

def find_label_row_in_block(ws, anchor_row: int, target_label: str = "ΕΠ.ΩΡΕΣ") -> int | None:
    target_norm = normalize_label(target_label)
    max_col = ws.max_column
    for r in range(anchor_row, anchor_row + 6):
        for c in range(max_col, 0, -1):
            v = ws.cell(row=r, column=c).value
            if normalize_label(v) == target_norm:
                return r
    return None

def get_epores_row(ws, rr: int) -> int:
    anchor = compute_anchor(rr)
    label_row = find_label_row_in_block(ws, anchor, "ΕΠ.ΩΡΕΣ")
    return label_row if label_row is not None else anchor

def normalize_repo_token(val):
    if val is None:
        return ""
    t = str(val).strip().upper()
    return "Ρ" if t == "Ρ" else t

def tag_schedule_rows_with_repo_from_form(
    schedule_rows,
    gui,
    forma_wb=None,
    forma_ws=None,
    start_row=10,
    end_row=1500,
    header_row=9,
    date_row=8,
    spreadsheet=None,
    month=None,
    get_column_from_day=None,
    strict_afm=True,
    write_guard=True
):
    import re
    from openpyxl.utils import column_index_from_string, get_column_letter

    AFM_COL_FORM = 1
    SUNDAY_COL_LETTER = "I"
    SUNDAY_COL = column_index_from_string(SUNDAY_COL_LETTER)

    if forma_ws is None:
        if forma_wb is None:
            gui.show_message("⛔ Δεν δόθηκε workbook ή φύλλο φόρμας", level="error")
            return schedule_rows
        forma_ws = forma_wb.active
        gui.show_message("⚠️ Χρήση ενεργού φύλλου ως φόρμα", level="warning")

    raw_date = forma_ws.cell(row=date_row, column=SUNDAY_COL).value
    gui.show_message(f"🗓️ Ανάγνωση ημερομηνίας Κυριακής από {SUNDAY_COL_LETTER}{date_row}: {raw_date!r} ({type(raw_date).__name__})", level="debug")
    try:
        if isinstance(raw_date, (datetime, date)):
            sunday_date = raw_date.date() if isinstance(raw_date, datetime) else raw_date
        else:
            sunday_date = datetime.strptime(str(raw_date), "%d/%m/%Y").date()
    except Exception:
        gui.show_message(f"⛔ Αδυναμία ανάγνωσης ημερομηνίας Κυριακής ➤ {SUNDAY_COL_LETTER}{date_row}: {raw_date}", level="error")
        return schedule_rows

    if month and sunday_date.month != int(month):
        gui.show_message(f"ℹ️ Κυριακή {sunday_date} εκτός στόχου μήνα ({month})", level="debug")
        return schedule_rows

    gui.show_message(f"📅 Ημερομηνία Κυριακής ➤ {sunday_date} (day={sunday_date.day})", level="debug")

    day_num = int(sunday_date.day)
    excel_col = None
    if get_column_from_day:
        try:
            excel_col = get_column_from_day(day_num)
            if isinstance(excel_col, int):
                excel_col = get_column_letter(excel_col)
            elif excel_col is not None:
                excel_col = str(excel_col).strip()
            gui.show_message(f"📌 Mapping ημέρας (provider) ➤ {day_num} → στήλη Excel: {excel_col!r}", level="debug")
        except Exception as e:
            gui.show_message(f"⛔ Αδυναμία εύρεσης στήλης από provider για ημέρα {day_num} ➤ {str(e)}", level="error")

    if not excel_col:
        col_index = 7 + day_num
        excel_col = get_column_letter(col_index)
        gui.show_message(f"🧭 Fallback mapping ημέρας ➤ {day_num} → {excel_col} (index={col_index})", level="debug")

    def to_date(d):
        try:
            return d.date() if hasattr(d, "date") else d
        except Exception:
            return d

    existing_map = {(str(e.get("employee")).strip(), to_date(e.get("date"))): e
                    for e in schedule_rows if isinstance(e, dict) and "employee" in e and "date" in e}

    orometrisi_ws = getattr(spreadsheet, "ws", None)
    if not orometrisi_ws:
        gui.show_message("⛔ Δεν υπάρχει φύλλο ΩΡΟΜΕΤΡΗΣΗ (spreadsheet.ws) για εγγραφή ΡΕΠΟ", level="error")

    # Build AFM -> rows mapping in ΩΡΟΜΕΤΡΗΣΗ once
    afm_to_rows = defaultdict(list)
    if orometrisi_ws:
        for rr in range(2, MAX_ROW + 1):
            v = orometrisi_ws.cell(row=rr, column=AFM_COL_ΩΡΟΜΕΤΡΗΣΗ).value
            if v is None:
                continue
            afm_to_rows[str(v).strip()].append(rr)

    added = updated = skipped = marked = not_found = overwritten = guarded = 0
    duplicate_afm_hits = 0

    gui.show_message(f"🔎 Σάρωση φόρμας για ΡΕΠΟ στη στήλη {SUNDAY_COL_LETTER}, γραμμές {start_row}..{end_row}", level="debug")

    seen_afms_written = set()

    def is_repo_from_form(val) -> bool:
        return normalize_label(val) == "ΡΕΠΟ"

    afms_with_repo = set()
    for r in range(start_row, end_row + 1):
        afm_raw = forma_ws.cell(row=r, column=AFM_COL_FORM).value
        if not afm_raw:
            continue
        afm = str(afm_raw).strip().split()[0]
        if strict_afm and not is_valid_afm(afm):
            continue
        val = forma_ws.cell(row=r, column=SUNDAY_COL).value
        if is_repo_from_form(val):
            afms_with_repo.add(afm)

    gui.show_message(f"📋 ΑΦΜ με ΡΕΠΟ στη φόρμα ({len(afms_with_repo)}): {sorted(afms_with_repo)}", level="debug")

    for r in range(start_row, end_row + 1):
        afm_raw = forma_ws.cell(row=r, column=AFM_COL_FORM).value
        if not afm_raw:
            continue

        afm = str(afm_raw).strip().split()[0]
        if strict_afm and not is_valid_afm(afm):
            skipped += 1
            gui.show_message(f"⏭️ Παράκαμψη (μη έγκυρο ΑΦΜ) ➤ '{afm_raw}' στη γραμμή {r}", level="debug")
            continue

        if afm not in afms_with_repo:
            skipped += 1
            continue

        if afm in seen_afms_written:
            skipped += 1
            gui.show_message(f"🛑 Παράκαμψη (ήδη γράφτηκε 'Ρ' σήμερα) ➤ ΑΦΜ {afm}", level="debug")
            continue

        key = (afm, sunday_date)
        if key in existing_map:
            e = existing_map[key]
            if not e.get("is_repo"):
                e["is_repo"] = True
                updated += 1
        else:
            schedule_rows.append({"date": sunday_date, "employee": afm, "is_repo": True})
            added += 1

        gui.show_message(f"➕ ΡΕΠΟ εντοπίστηκε στη φόρμα ➤ ΑΦΜ {afm} | {sunday_date.strftime('%d/%m/%Y')} (row {r})", level="debug")

        if not orometrisi_ws:
            continue

        afm_clean = str(afm).strip()
        match_rows = afm_to_rows.get(afm_clean, [])

        if not match_rows:
            not_found += 1
            gui.show_message(f"⚠️ Δεν βρέθηκε ΑΦΜ {afm_clean} στο ΩΡΟΜΕΤΡΗΣΗ", level="debug")
            continue

        if len(match_rows) > 1:
            duplicate_afm_hits += 1
            gui.show_message(f"⚠️ Πολλαπλές εμφανίσεις ΑΦΜ {afm_clean} στο ΩΡΟΜΕΤΡΗΣΗ ➤ {match_rows}", level="debug")

        wrote_for_this_afm = False
        for rr in match_rows:
            if afm_clean in seen_afms_written:
                break
            anchor = compute_anchor(rr)
            target_row = get_epores_row(orometrisi_ws, rr)
            target_col = column_index_from_string(excel_col)
            cell_a1 = f"{excel_col}{target_row}"

            if afm_clean not in afms_with_repo:
                gui.show_message(f"⛔ Ασυμφωνία: Απόπειρα εγγραφής 'Ρ' για ΑΦΜ {afm_clean} χωρίς ΡΕΠΟ στη φόρμα", level="error")
                break

            if write_guard:
                existing = orometrisi_ws.cell(row=target_row, column=target_col).value
                if normalize_repo_token(existing) == "Ρ":
                    guarded += 1
                    gui.show_message(f"🛡️ Παράκαμψη εγγραφής ➤ {cell_a1} έχει ήδη 'Ρ'", level="debug")
                    continue
                if existing not in (None, "") and normalize_repo_token(existing) != "Ρ":
                    overwritten += 1
                    gui.show_message(f"⚠️ Overwrite ➤ {cell_a1}: {existing!r} → 'Ρ'", level="warning")

            before = orometrisi_ws[cell_a1].value
            orometrisi_ws[cell_a1].value = "Ρ"
            after = orometrisi_ws[cell_a1].value
            gui.show_message(f"✏️ Εγγραφή ΡΕΠΟ στο {cell_a1} ➤ πριν: {before!r} → μετά: {after!r}", level="debug")

            if hasattr(spreadsheet, "update_cell"):
                try:
                    spreadsheet.update_cell(cell_a1, "Ρ")
                except Exception:
                    pass

            marked += 1
            seen_afms_written.add(afm_clean)
            wrote_for_this_afm = True
            break

        if not wrote_for_this_afm:
            gui.show_message(f"ℹ️ Δεν πραγματοποιήθηκε εγγραφή 'Ρ' για ΑΦΜ {afm_clean} (guards/anchors/matches)", level="debug")

    gui.show_message(f"📌 Γράφτηκε 'Ρ' για ΑΦΜ: {sorted(seen_afms_written)}", level="debug")
    gui.show_message(
        f"✅ Ολοκλήρωση: προστέθηκαν={added}, ενημερώθηκαν={updated}, "
        f"γράφτηκαν={marked}, παρακάμφθηκαν={guarded}, δεν βρέθηκαν={not_found}, πολλαπλά={duplicate_afm_hits}",
        level="info"
    )
    return schedule_rows

def generate_monthly_report(
    schedule_rows,
    month,
    spreadsheet,
    gui,
    get_column_from_day,
    overtime_ws=None,
    forma_wb=None,
    forma_ws=None
):
    from datetime import datetime
    from calendar import monthrange

    updated_count = 0
    skipped_count = 0
    repo_entries = 0
    processed_entries = 0
    all_row_lists = []

    if spreadsheet is None or getattr(spreadsheet, "ws", None) is None:
        gui.show_message("⛔ generate_monthly_report: Δεν υπάρχει διαθέσιμο φύλλο ΩΡΟΜΕΤΡΗΣΗ (spreadsheet.ws)", level="error")
        return 0, 0

    ws_orometrisi = spreadsheet.ws

    if overtime_ws is None:
        try:
            wb = getattr(spreadsheet, "wb", None)
            if wb and "ΥΠΕΡΕΡΓΑΣΙΕΣ-ΥΠΕΡΩΡΙΕΣ" in wb.sheetnames:
                overtime_ws = wb["ΥΠΕΡΕΡΓΑΣΙΕΣ-ΥΠΕΡΩΡΙΕΣ"]
                gui.show_message("📄 Χρήση φύλλου 'ΥΠΕΡΕΡΓΑΣΙΕΣ-ΥΠΕΡΩΡΙΕΣ' ως πηγή ωρών", level="debug")
            else:
                gui.show_message("⛔ Δεν βρέθηκε φύλλο 'ΥΠΕΡΕΡΓΑΣΙΕΣ-ΥΠΕΡΩΡΙΕΣ'", level="error")
        except Exception as ex:
            gui.show_message(f"⛔ Αποτυχία πρόσβασης σε workbook: {ex}", level="error")

    sample_year = schedule_rows[0]["date"].year if schedule_rows else datetime.now().year
    max_day = monthrange(sample_year, month)[1]
    gui.show_message(f"📅 Ο μήνας {month} του {sample_year} έχει {max_day} ημέρες", level="debug")

    day_to_excel_col = {}
    for d in range(1, max_day + 1):
        try:
            col = get_column_from_day(d)
            day_to_excel_col[d] = col
        except Exception:
            day_to_excel_col[d] = None

    gui.show_message("🏷️ Εκκίνηση tagging ΡΕΠΟ από ΦΟΡΜΑ", level="debug")
    schedule_rows = tag_schedule_rows_with_repo_from_form(
        schedule_rows=schedule_rows,
        gui=gui,
        forma_wb=forma_wb,
        forma_ws=forma_ws,
        start_row=10,
        end_row=150,
        header_row=9,
        date_row=8,
        spreadsheet=spreadsheet,
        month=month,
        get_column_from_day=get_column_from_day,
        strict_afm=True,
        write_guard=True
    )
    gui.show_message("🏁 Ολοκλήρωση tagging ΡΕΠΟ από ΦΟΡΜΑ", level="debug")

    repo_entries = sum(1 for e in schedule_rows if e.get("is_repo"))
    total_entries = len(schedule_rows)
    gui.show_message(f"🧮 Σύνοψη schedule_rows ➤ σύνολο={total_entries}, με ΡΕΠΟ={repo_entries}", level="debug")

    # Cache is sheet-aware now: keys are (id(ws), afm)
    afm_cache = {}

    for idx, entry in enumerate(schedule_rows, start=1):
        processed_entries += 1
        date_obj = entry["date"]
        afm = entry["employee"]
        hours = entry.get("hours")
        work_type = (entry.get("work_type") or "").strip().upper()

        gui.show_message(
            f"📄 [{idx}/{total_entries}] Επεξεργασία ➤ ΑΦΜ: {afm}, ημερομηνία: {date_obj}, ώρες: {hours}, τύπος: {work_type}, repo={entry.get('is_repo', False)}",
            level="debug"
        )

        if date_obj.month != month:
            gui.show_message(f"⏩ Παράκαμψη μήνα ➤ {date_obj.month} ≠ {month}", level="debug")
            continue
        if date_obj.day > max_day:
            gui.show_message(f"⚠️ Ημέρα {date_obj.day} υπερβαίνει τις {max_day}", level="warning")
            skipped_count += 1
            continue

        if entry.get("is_repo", False):
            if date_obj.weekday() != 6:
                gui.show_message(f"ℹ️ (ΡΕΠΟ) Ημέρα {date_obj} δεν είναι Κυριακή → Καμία ενέργεια", level="debug")
                continue

            row_list = find_employee_row_in_sheet(ws_orometrisi, afm, gui=gui, diagnostics=True, cache=afm_cache)
            if not row_list:
                gui.show_message(f"⚠️ (ΡΕΠΟ) Δεν βρέθηκε εργαζόμενος στο φύλλο ➤ {afm}", level="warning")
                skipped_count += 1
                continue

            excel_col = day_to_excel_col.get(date_obj.day) or get_column_from_day(date_obj.day)
            metric_rows = get_metric_rows(ws_orometrisi, row_list[0])

            updated_count += 1
            continue

        if work_type == "6ΗΜΕΡΟΣ":
            base_hours = 6.67
        elif work_type == "5ΗΜΕΡΟΣ":
            base_hours = 8.0
        else:
            base_hours = 0

        row_list = find_employee_row_in_sheet(ws_orometrisi, afm, gui=gui, diagnostics=True, cache=afm_cache)
        if not row_list:
            gui.show_message(f"⚠️ Δεν βρέθηκε εργαζόμενος στο φύλλο ➤ {afm}", level="warning")
            skipped_count += 1
            continue

        all_row_lists.append(row_list)
        excel_col = day_to_excel_col.get(date_obj.day) or get_column_from_day(date_obj.day)

        if overtime_ws:
            overtime_anchor_list = find_employee_row_in_sheet(overtime_ws, afm, gui=gui, diagnostics=True, cache=afm_cache)
            overtime_anchor = overtime_anchor_list[0] if overtime_anchor_list else 0
        else:
            overtime_anchor = 0

        if not overtime_anchor:
            gui.show_message(f"⚠️ Δεν βρέθηκε anchor στο φύλλο ωρών για {afm}", level="warning")
            continue

        times = read_work_times_from_sheet(overtime_ws, overtime_anchor, date_obj, gui=gui)
        raw_end_plus_30 = times.get("ΩΡΑ ΛΗΞΗΣ+30")
        raw_departure = times.get("ΩΡΑ ΑΠΟΧΩΡΗΣΗ")

        end_plus_30 = clean_time_string(raw_end_plus_30)
        departure_time = clean_time_string(raw_departure)

        if not is_valid_time_string(departure_time):
            if date_obj.weekday() == 6:
                gui.show_message(
                    f"📅 Κυριακή χωρίς αποχώρηση ➤ '{raw_departure}' → Καταγραφή ως ΑΡΓΙΑ (base_hours)",
                    level="debug"
                )
                metric_rows = get_metric_rows(ws_orometrisi, row_list[0])

                if "ΑΡΓΙΑ" in metric_rows:
                    cell_name = f"{excel_col}{metric_rows['ΑΡΓΙΑ']}"
                    update_excel_cell(ws_orometrisi, cell_name, round(base_hours, 2))
                    gui.show_message(f"🧾 ΑΡΓΙΑ ➤ {cell_name} ➤ {round(base_hours, 2)}", level="debug")

                if "ΠΛΗΘΟΣ ΚΥΡΙΑΚΩΝ" in metric_rows:
                    cell_name = f"{excel_col}{metric_rows['ΠΛΗΘΟΣ ΚΥΡΙΑΚΩΝ']}"
                    update_excel_cell(ws_orometrisi, cell_name, 1)
                    gui.show_message(f"📅 ΠΛΗΘΟΣ ΚΥΡΙΑΚΩΝ ➤ {cell_name} ➤ 1", level="debug")

                updated_count += 1
                continue
            else:
                gui.show_message(f"⏭️ Δεν υπάρχει αποχώρηση ➤ '{raw_departure}' → Δεν υπολογίζεται υπερωρία", level="debug")
                continue

        if not is_valid_time_string(end_plus_30):
            gui.show_message(f"⚠️ Μη έγκυρη ώρα λήξης+30 ➤ '{raw_end_plus_30}' → Παράκαμψη", level="warning")
            continue

        gui.show_message(f"⏱️ Υπολογισμός υπερωριών ➤ Λήξη+30': {end_plus_30}, Αποχώρηση: {departure_time}", level="debug")
        results = calculate_overtime(end_plus_30, departure_time, date_obj)
        gui.show_message(f"📊 Αποτελέσματα ➤ Υπερεργασία: {results['ΥΠΕΡΕΡΓΑΣΙΑ']}, Υπερωρία: {results['ΥΠΕΡΩΡΙΑ']}, Αργία: {results['ΑΡΓΙΑ']}", level="debug")

        metric_rows = get_metric_rows(ws_orometrisi, row_list[0])

        if "ΑΡΓΙΑ" in metric_rows:
            if date_obj.weekday() == 6:
                total_argia = round((6.67 if work_type == "6ΗΜΕΡΟΣ" else 8.0 if work_type == "5ΗΜΕΡΟΣ" else 0) + float(results.get("ΑΡΓΙΑ", 0)), 2)
                if total_argia > 0:
                    cell_name = f"{excel_col}{metric_rows['ΑΡΓΙΑ']}"
                    update_excel_cell(ws_orometrisi, cell_name, total_argia)
                    gui.show_message(f"🧾 ΑΡΓΙΑ ➤ {cell_name} ➤ {total_argia} (base + υπερεργασία + υπερωρία)", level="debug")
            else:
                if results["ΑΡΓΙΑ"] > 0:
                    cell_name = f"{excel_col}{metric_rows['ΑΡΓΙΑ']}"
                    update_excel_cell(ws_orometrisi, cell_name, results["ΑΡΓΙΑ"])
                    gui.show_message(f"🧾 ΑΡΓΙΑ ➤ {cell_name} ➤ {results['ΑΡΓΙΑ']}", level="debug")

        if results["ΥΠΕΡΕΡΓΑΣΙΑ"] > 0 and "ΥΠΕΡΕΡΓΑΣΙΑ" in metric_rows:
            cell_name = f"{excel_col}{metric_rows['ΥΠΕΡΕΡΓΑΣΙΑ']}"
            update_excel_cell(ws_orometrisi, cell_name, results["ΥΠΕΡΕΡΓΑΣΙΑ"])
            gui.show_message(f"🧾 ΥΠΕΡΕΡΓΑΣΙΑ ➤ {cell_name} ➤ {results['ΥΠΕΡΕΡΓΑΣΙΑ']}", level="debug")

        if results["ΥΠΕΡΩΡΙΑ"] > 0 and "ΥΠΕΡΩΡΙΑ" in metric_rows:
            cell_name = f"{excel_col}{metric_rows['ΥΠΕΡΩΡΙΑ']}"
            update_excel_cell(ws_orometrisi, cell_name, results["ΥΠΕΡΩΡΙΑ"])
            gui.show_message(f"🧾 ΥΠΕΡΩΡΙΑ ➤ {cell_name} ➤ {results['ΥΠΕΡΩΡΙΑ']}", level="debug")

        if date_obj.weekday() == 6 and "ΠΛΗΘΟΣ ΚΥΡΙΑΚΩΝ" in metric_rows:
            cell_name = f"{excel_col}{metric_rows['ΠΛΗΘΟΣ ΚΥΡΙΑΚΩΝ']}"
            update_excel_cell(ws_orometrisi, cell_name, 1)
            gui.show_message(f"📅 ΠΛΗΘΟΣ ΚΥΡΙΑΚΩΝ ➤ {cell_name} ➤ 1", level="debug")

        night_hours = calculate_night_hours(end_plus_30, departure_time)
        gui.show_message(f"🌒 Νυχτερινό ➤ {night_hours} ώρες (από {end_plus_30} έως {departure_time})", level="debug")

        if night_hours > 0 and "ΝΥΧΤΑ" in metric_rows:
            cell_name = f"{excel_col}{metric_rows['ΝΥΧΤΑ']}"
            update_excel_cell(ws_orometrisi, cell_name, night_hours)
            gui.show_message(f"🧾 ΝΥΧΤΑ ➤ {cell_name} ➤ {night_hours}", level="debug")

        updated_count += 1

    gui.show_message(
        f"✅ Ολοκλήρωση ➤ Ενημερώθηκαν {updated_count} εγγραφές, παρακάμφθηκαν {skipped_count} | "
        f"σύνολο processed={processed_entries}, με ΡΕΠΟ={repo_entries}",
        level="info"
    )
    return updated_count, skipped_count